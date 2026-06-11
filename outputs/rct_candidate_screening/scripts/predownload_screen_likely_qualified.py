import csv
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROCESSED = ROOT / "processed"
OUT = ROOT / "predownload_screen"
OUT.mkdir(parents=True, exist_ok=True)

CANDIDATES = PROCESSED / "candidate_records.csv"
EXPANSION_DOWNLOAD_LOG = Path("rct_expansion/provenance/download_log.csv")
ORIGINAL_META = Path("cleaned_data/meta_data.xlsx")
FLOW_COUNTS = Path("rct_expansion/provenance/broad_dataset_screening_flow_counts.csv")

METADATA_FIELDS = [
    "title",
    "publisher",
    "repository_client",
    "provider",
    "doi",
    "url",
    "creators",
    "formats",
    "rights_license_short",
    "subjects_keywords_short",
    "related_publications_short",
    "related_identifiers_short",
    "description_short",
]

DIRECT_REPOSITORY_CLIENTS = {
    "dryad.dryad",
    "gdcc.harvard-dv",
    "cern.zenodo",
    "bl.mendeley",
    "dans.dataversenl",
    "ocul.spdv",
    "bl.lboro",
    "bl.bath",
    "figshare.plus",
}

TABULAR_OR_DATA_FILE_RE = re.compile(
    r"csv|excel|spreadsheet|stata|spss|tab-separated|text/tab|sav|dta|xlsx|xls|tsv|\.tab\b|application/zip|text/plain",
    re.I,
)
PUBLICATION_DOI_RE = re.compile(r"\b10\.\d{4,9}/[^\s|;,)]*", re.I)
RESTRICTION_RE = re.compile(
    r"restrictedaccess|restricted access|closedaccess|closed access|custom terms|data use agreement|\bdua\b|request access|available upon request|provided on request|on request|controlled access|login required|cc-by-nc|non commercial|no derivatives|permission required|embargo",
    re.I,
)
OPEN_LICENSE_RE = re.compile(r"cc0|cc-by|creative commons|openaccess|open access|public domain", re.I)
HEALTH_CLINICAL_RE = re.compile(
    r"\b(patient|patients|participant|participants|clinical|clinic|medical|medicine|health|disease|pain|surgery|therapy|treatment|drug|dose|pharma|hospital|diabetes|cancer|covid|stroke|depression|anxiety|psych|obesity|weight|cardiac|renal|kidney|pregnan|postpartum|dermatitis|nausea|vomiting|atrial|fibrillation|adults|children|women|men|older|elderly|neuro|rehabilitation|symptom|blood|infection|hiv|malaria|vaccin|exercise|diet|sleep|back pain|bmi|glycemia|insulin)\b",
    re.I,
)
PARTICIPANT_RE = re.compile(
    r"\b(patient|patients|participant|participants|adults|children|students|workers|women|men|placebo|treatment|intervention|control|assigned|allocated|randomly assigned|randomly allocated)\b",
    re.I,
)
ANIMAL_RE = re.compile(
    r"\b(calf|calves|cat|cats|dog|dogs|mouse|mice|rat|rats|animal|veterinary|dairy|shelter cats|cattle|bovine|porcine|swine|sheep|goat|poultry)\b",
    re.I,
)
OBSERVATIONAL_RE = re.compile(
    r"cross-sectional|observational|cohort study|case-control|survey of|attitudes, willingness|knowledge, attitudes|secondary analysis|interim analysis|pooled analysis|retrospective",
    re.I,
)
PROTOCOL_OR_MATERIAL_RE = re.compile(
    r"study protocol|trial protocol|protocol for|statistical analysis plan|questionnaire|instrument|codebook only|materials only",
    re.I,
)
CLUSTER_RE = re.compile(
    r"cluster[- ]randomi[sz]ed|community[- ]randomi[sz]ed|school[- ]randomi[sz]ed|stepped wedge",
    re.I,
)
AGGREGATE_OR_SIMULATION_RE = re.compile(
    r"aggregate|aggregated|summary-level|summary data|simulation|simulated|manikin|cadaver",
    re.I,
)
META_OR_NON_CLINICAL_RE = re.compile(
    r"case report|meta-analysis|systematic review|whole genome|genome analysis|microfinance|public goods|business training|entrepreneurship|policy experiment|development economics|gendered attitudes|women working|early childhood education|school gardens|school and home gardens|study within a trial|\\bswat\\b|trial recruitment|participant testimony|online survey|qualtrics|private providers|retail drug outlets",
    re.I,
)
CODE_ONLY_TITLE_RE = re.compile(r"\breplication code\b|\bcode for\b", re.I)


def norm_doi(value):
    value = (value or "").strip().lower()
    value = value.replace("https://doi.org/", "").replace("http://doi.org/", "")
    value = value.replace("doi:", "").strip().strip(".").strip("/")
    return value


def doi_tokens(text):
    return {norm_doi(x) for x in PUBLICATION_DOI_RE.findall(text or "") if norm_doi(x)}


def metadata_text(row):
    return " ".join((row.get(field) or "") for field in METADATA_FIELDS).lower()


def load_active_expansion_dois():
    dois = set()
    if not EXPANSION_DOWNLOAD_LOG.exists():
        return dois
    with open(EXPANSION_DOWNLOAD_LOG, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            for token in re.split(r";|\s+", row.get("Dataset_DOI", "")):
                token = norm_doi(token)
                if token.startswith("10."):
                    dois.add(token)
    return dois


def load_original_paper_dois():
    dois = set()
    if not ORIGINAL_META.exists():
        return dois
    wb = load_workbook(ORIGINAL_META, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        dois.update(doi_tokens(str(row.get("Paper Link") or "")))
    return dois


def candidate_dois(row):
    text = " ".join(
        [
            row.get("doi", ""),
            row.get("related_publications_short", ""),
            row.get("related_identifiers_short", ""),
        ]
    )
    dois = doi_tokens(text)
    if row.get("doi"):
        dois.add(norm_doi(row["doi"]))
    # Normalize Zenodo concept DOI/version sibling matches.
    expanded = set(dois)
    for doi in list(dois):
        if doi.startswith("10.5281/zenodo."):
            expanded.add(re.sub(r"\.\d+$", "", doi))
    return expanded


def screen(row, active_dois, original_paper_dois):
    text = metadata_text(row)
    title = row.get("title", "")
    dois = candidate_dois(row)
    active_hits = sorted(active_dois & dois)
    original_hits = sorted(original_paper_dois & dois)

    direct_repository = row.get("repository_client") in DIRECT_REPOSITORY_CLIENTS
    explicit_file = bool(row.get("file_count")) or bool(
        TABULAR_OR_DATA_FILE_RE.search((row.get("formats") or "") + " " + (row.get("description_short") or ""))
    )
    publication_doi = bool(
        doi_tokens((row.get("related_publications_short") or "") + " " + (row.get("related_identifiers_short") or ""))
    )
    health_signal = bool(HEALTH_CLINICAL_RE.search(text))
    participant_signal = bool(PARTICIPANT_RE.search(text))
    open_license_signal = bool(OPEN_LICENSE_RE.search(row.get("rights_license_short") or ""))
    restriction_signal = bool(RESTRICTION_RE.search(text))

    red_flags = []
    if active_hits:
        red_flags.append("already active expansion DOI")
    if original_hits:
        red_flags.append("matches original benchmark paper DOI")
    if restriction_signal:
        red_flags.append("restriction/custom/DUA signal")
    if ANIMAL_RE.search(text):
        red_flags.append("animal/veterinary signal")
    if OBSERVATIONAL_RE.search(text):
        red_flags.append("observational/cross-sectional/secondary-analysis signal")
    if PROTOCOL_OR_MATERIAL_RE.search(text):
        red_flags.append("protocol/material/instrument signal")
    if CLUSTER_RE.search(text):
        red_flags.append("cluster-randomized/stepped-wedge signal")
    if AGGREGATE_OR_SIMULATION_RE.search(text):
        red_flags.append("aggregate/simulation/manikin signal")
    if META_OR_NON_CLINICAL_RE.search(text):
        red_flags.append("meta-analysis/genome/non-clinical policy signal")
    if CODE_ONLY_TITLE_RE.search(title):
        red_flags.append("replication-code-only title signal")

    score = 0
    score += 3 if direct_repository else 0
    score += 3 if explicit_file else 0
    score += 2 if publication_doi else 0
    score += 2 if health_signal else 0
    score += 1 if participant_signal else 0
    score += 1 if open_license_signal else 0
    score -= 8 if active_hits else 0
    score -= 8 if original_hits else 0
    score -= 5 if restriction_signal else 0
    score -= 4 * max(0, len(red_flags) - bool(active_hits) - bool(original_hits) - bool(restriction_signal))

    if active_hits:
        status = "Already active expansion - exclude"
    elif original_hits:
        status = "Already original benchmark - exclude"
    elif red_flags:
        status = "Exclude before download"
    elif direct_repository and explicit_file and publication_doi and health_signal:
        status = "Download first"
    elif direct_repository and publication_doi and health_signal:
        status = "Landing page / file manifest first"
    elif explicit_file and publication_doi and health_signal:
        status = "Source-specific manual check"
    elif publication_doi and health_signal:
        status = "Defer - missing access/file certainty"
    else:
        status = "Defer - low clinical or data confidence"

    signals = []
    if direct_repository:
        signals.append("direct/open repository client")
    if explicit_file:
        signals.append("explicit file/tabular signal")
    if publication_doi:
        signals.append("publication DOI signal")
    if health_signal:
        signals.append("human clinical/health signal")
    if participant_signal:
        signals.append("participant/treatment-arm signal")
    if open_license_signal:
        signals.append("open license signal")
    if active_hits:
        signals.append("active expansion DOI: " + "; ".join(active_hits))
    if original_hits:
        signals.append("original paper DOI: " + "; ".join(original_hits))
    signals.extend(red_flags)

    order = {
        "Download first": 1,
        "Landing page / file manifest first": 2,
        "Source-specific manual check": 3,
        "Defer - missing access/file certainty": 4,
        "Defer - low clinical or data confidence": 5,
        "Already active expansion - exclude": 6,
        "Already original benchmark - exclude": 7,
        "Exclude before download": 8,
    }[status]

    return {
        "predownload_status": status,
        "predownload_order": order,
        "predownload_score": score,
        "direct_repository_signal": "yes" if direct_repository else "no",
        "explicit_file_signal": "yes" if explicit_file else "no",
        "publication_doi_signal": "yes" if publication_doi else "no",
        "human_clinical_health_signal": "yes" if health_signal else "no",
        "participant_signal": "yes" if participant_signal else "no",
        "open_license_signal": "yes" if open_license_signal else "no",
        "restriction_signal_strict": "yes" if restriction_signal else "no",
        "active_expansion_doi_match": "; ".join(active_hits),
        "original_paper_doi_match": "; ".join(original_hits),
        "predownload_reasons": "; ".join(dict.fromkeys(signals)),
    }


def write_csv(path, rows, fields):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main():
    active_dois = load_active_expansion_dois()
    original_paper_dois = load_original_paper_dois()
    with open(CANDIDATES, newline="", encoding="utf-8") as f:
        likely_rows = [row for row in csv.DictReader(f) if row.get("screen_status") == "Likely qualified"]

    rows = []
    for row in likely_rows:
        row = dict(row)
        row.update(screen(row, active_dois, original_paper_dois))
        rows.append(row)

    rows.sort(
        key=lambda r: (
            int(r["predownload_order"]),
            -int(r["predownload_score"]),
            r.get("publisher", ""),
            r.get("title", ""),
        )
    )

    screen_fields = [
        "candidate_id",
        "predownload_status",
        "predownload_score",
        "predownload_reasons",
        "direct_repository_signal",
        "explicit_file_signal",
        "publication_doi_signal",
        "human_clinical_health_signal",
        "participant_signal",
        "open_license_signal",
        "restriction_signal_strict",
        "active_expansion_doi_match",
        "original_paper_doi_match",
        "title",
        "publisher",
        "repository_client",
        "doi",
        "url",
        "publication_year",
        "formats",
        "rights_license_short",
        "related_publications_short",
        "description_short",
    ]
    write_csv(OUT / "likely_qualified_predownload_screen.csv", rows, screen_fields)
    write_csv(
        OUT / "download_first_shortlist.csv",
        [r for r in rows if r["predownload_status"] == "Download first"],
        screen_fields,
    )
    write_csv(
        OUT / "landing_page_first_shortlist.csv",
        [r for r in rows if r["predownload_status"] == "Landing page / file manifest first"],
        screen_fields,
    )

    counts = Counter(r["predownload_status"] for r in rows)
    summary = [
        {"summary_type": "predownload_status", "group": k, "count": counts[k]}
        for k in sorted(counts, key=lambda status: min(int(r["predownload_order"]) for r in rows if r["predownload_status"] == status))
    ]
    for status in counts:
        by_pub = Counter(r["publisher"] or "Unknown" for r in rows if r["predownload_status"] == status)
        for publisher, count in by_pub.most_common(15):
            summary.append(
                {
                    "summary_type": f"top_publishers | {status}",
                    "group": publisher,
                    "count": count,
                }
            )
    write_csv(OUT / "predownload_summary.csv", summary, ["summary_type", "group", "count"])

    rules = [
        {
            "rule": "Already curated",
            "logic": "Exclude before new work when candidate DOI or related publication DOI matches active expansion download_log.csv or original meta_data.xlsx paper DOI.",
        },
        {
            "rule": "Download first",
            "logic": "Direct repository client, explicit file/tabular signal, publication DOI signal, human clinical/health signal, and no strict red flags.",
        },
        {
            "rule": "Landing page / file manifest first",
            "logic": "Direct repository client plus publication DOI and health signal, but missing explicit file evidence. Inspect landing page or file manifest only before downloading data.",
        },
        {
            "rule": "Source-specific manual check",
            "logic": "Explicit data-file signal, publication DOI, and health signal, but source is not a preferred direct open repository or may need special access/source interpretation.",
        },
        {
            "rule": "Defer - missing access/file certainty",
            "logic": "Publication-backed health record, but source or file/access metadata is too sparse for a first download batch.",
        },
        {
            "rule": "Defer - low clinical or data confidence",
            "logic": "Metadata does not strongly support human clinical/health RCT data or publication/file readiness.",
        },
        {
            "rule": "Exclude before download",
            "logic": "Restriction/custom/DUA, animal/veterinary, observational/secondary analysis, protocol/materials-only, cluster/stepped-wedge, aggregate/simulation, meta-analysis/genome-only, or non-clinical policy/economics signal.",
        },
    ]
    write_csv(OUT / "predownload_rules.csv", rules, ["rule", "logic"])

    stats = {
        "input_likely_qualified": len(rows),
        "counts": dict(counts),
        "active_expansion_dois_loaded": len(active_dois),
        "original_paper_dois_loaded": len(original_paper_dois),
    }
    (OUT / "predownload_stats.json").write_text(json.dumps(stats, indent=2), encoding="utf-8")
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
