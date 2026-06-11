#!/usr/bin/env python3
"""Screen manually downloaded RCTC files for benchmark expansion.

This script inventories the files dropped in manually-download-raw-data, extracts
archives into this provenance folder, and writes compact table/schema summaries
used for qualification decisions. It deliberately avoids writing active cleaned
trial outputs.
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import subprocess
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependency availability is recorded.
    load_workbook = None

try:
    import pyreadstat
except Exception:  # pragma: no cover - dependency availability is recorded.
    pyreadstat = None


ROOT = Path(__file__).resolve().parents[3]
WORKSPACE = Path(__file__).resolve().parent
MANUAL_DIR = ROOT / "manually-download-raw-data"
EXTRACTED_DIR = WORKSPACE / "extracted"
SCREENING_RECORDS = ROOT / "outputs/rct_candidate_screening/processed/candidate_records.csv"
PREDOWNLOAD = ROOT / "outputs/rct_candidate_screening/predownload_screen/likely_qualified_predownload_screen.csv"
MAX_ROWS = 5000


def candidate_id_from_text(text: str) -> str | None:
    text = text.replace("RCRC", "RCTC")
    if re.search(r"(^|/)Raw_data\.xlsx$", text) or re.search(r"(^|/)read_me_file\.docx$", text):
        return "RCTC-01363"
    match = re.search(r"RCTC[-_ ]?0*(\d{2,5})", text, flags=re.I)
    if not match:
        return None
    return f"RCTC-{int(match.group(1)):05d}"


def load_candidate_metadata() -> dict[str, dict[str, str]]:
    rows: dict[str, dict[str, str]] = {}
    for path in [SCREENING_RECORDS, PREDOWNLOAD]:
        if not path.exists():
            continue
        with path.open(newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                cid = row.get("candidate_id") or row.get("Candidate ID") or ""
                if cid and cid not in rows:
                    rows[cid] = row
    return rows


def copy_or_extract_sources() -> list[Path]:
    if EXTRACTED_DIR.exists():
        shutil.rmtree(EXTRACTED_DIR)
    EXTRACTED_DIR.mkdir(parents=True, exist_ok=True)

    paths: list[Path] = []
    for path in MANUAL_DIR.rglob("*"):
        if not path.is_file() or path.name.startswith("."):
            continue
        cid = candidate_id_from_text(str(path.relative_to(MANUAL_DIR)))
        if not cid:
            cid = "UNMAPPED"
        if path.suffix.lower() == ".zip":
            target = EXTRACTED_DIR / cid / path.stem
            target.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(path) as zf:
                    zf.extractall(target)
            except Exception as exc:
                (target / "_extract_error.txt").write_text(str(exc), encoding="utf-8")
            continue
        paths.append(path)

    for path in EXTRACTED_DIR.rglob("*"):
        if path.is_file() and not path.name.startswith("."):
            paths.append(path)
    return sorted(paths)


def safe_rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def file_role(path: Path) -> str:
    ext = path.suffix.lower()
    if path.name == "RCRC-00890":
        return "data_or_table"
    if ext in {".csv", ".tsv", ".tab", ".txt", ".xlsx", ".xls", ".ods", ".sav", ".dta"}:
        return "data_or_table"
    if ext in {".md", ".docx", ".pdf"}:
        return "support"
    return "other"


def read_delimited(path: Path) -> list[dict[str, object]]:
    sep = "\t" if path.suffix.lower() in {".tsv", ".tab"} else None
    try:
        df = pd.read_csv(path, sep=sep, nrows=MAX_ROWS, engine="python", encoding="utf-8")
    except Exception:
        try:
            df = pd.read_csv(path, sep=sep, nrows=MAX_ROWS, engine="python", encoding="latin1")
        except Exception as exc:
            return [{"sheet": "", "read_status": "unreadable", "error": str(exc)[:250]}]
    return [summarize_df(df, path.name, "delimited", sampled_rows=len(df))]


def read_xlsx(path: Path) -> list[dict[str, object]]:
    if load_workbook is None:
        return [{"sheet": "", "read_status": "dependency_missing", "error": "openpyxl unavailable"}]
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [{"sheet": "", "read_status": "unreadable", "error": str(exc)[:250]}]
    out = []
    for sheet in wb.sheetnames[:20]:
        ws = wb[sheet]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(row))
            if i >= MAX_ROWS:
                break
        if not rows:
            out.append({"sheet": sheet, "read_status": "empty", "error": ""})
            continue
        header = make_unique_names([str(x).strip() if x is not None else f"X{i+1}" for i, x in enumerate(rows[0])])
        body = rows[1:]
        df = pd.DataFrame(body, columns=header)
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        out.append(summarize_df(df, path.name, sheet, sampled_rows=len(df)))
    return out


def read_excel_with_pandas(path: Path) -> list[dict[str, object]]:
    try:
        book = pd.ExcelFile(path)
    except Exception as exc:
        return [{"sheet": "", "read_status": "unreadable", "error": str(exc)[:250]}]
    out = []
    for sheet in book.sheet_names[:20]:
        try:
            df = pd.read_excel(book, sheet_name=sheet, nrows=MAX_ROWS)
        except Exception as exc:
            out.append({"sheet": sheet, "read_status": "unreadable", "error": str(exc)[:250]})
            continue
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        out.append(summarize_df(df, path.name, sheet, sampled_rows=len(df)))
    return out


def make_unique_names(names: list[str]) -> list[str]:
    seen: defaultdict[str, int] = defaultdict(int)
    out = []
    for i, name in enumerate(names):
        clean = name if name else f"X{i+1}"
        seen[clean] += 1
        out.append(clean if seen[clean] == 1 else f"{clean}_{seen[clean]}")
    return out


def read_foreign(path: Path) -> list[dict[str, object]]:
    if pyreadstat is not None:
        try:
            if path.suffix.lower() == ".sav":
                df, _ = pyreadstat.read_sav(str(path))
            else:
                df, _ = pyreadstat.read_dta(str(path))
            return [summarize_df(df, path.name, path.suffix.lower().lstrip(".") or "stata", sampled_rows=len(df))]
        except Exception:
            pass

    script = WORKSPACE / "_foreign_summary.R"
    script.write_text(
        """
args <- commandArgs(trailingOnly=TRUE)
path <- args[[1]]
ext <- tolower(tools::file_ext(path))
d <- NULL
err <- ""
tryCatch({
  if (ext == "sav") {
    d <- foreign::read.spss(path, to.data.frame=TRUE, use.value.labels=FALSE, trim.factor.names=TRUE)
  } else if (ext == "dta") {
    d <- foreign::read.dta(path, convert.factors=FALSE)
  }
}, error=function(e) { err <<- conditionMessage(e) })
if (!identical(err, "") || is.null(d)) {
  cat("ERROR\\t", substr(gsub("[\\r\\n\\t]+", " ", err), 1, 250), "\\n", sep="")
} else {
  nms <- names(d)
  miss <- sapply(d, function(x) mean(is.na(x)))
  uniq <- sapply(d, function(x) length(unique(x[!is.na(x)])))
  head_vals <- sapply(d, function(x) paste(head(unique(as.character(x[!is.na(x)])), 4), collapse="|"))
  clean <- function(x) gsub("[\\r\\n\\t]+", " ", as.character(x))
  cat("OK\\t", nrow(d), "\\t", ncol(d), "\\n", sep="")
  for (i in seq_along(nms)) {
    cat(clean(nms[[i]]), "\\t", sprintf("%.3f", miss[[i]]), "\\t", uniq[[i]], "\\t", clean(head_vals[[i]]), "\\n", sep="")
  }
}
""",
        encoding="utf-8",
    )
    try:
        proc = subprocess.run(["Rscript", str(script), str(path)], text=True, capture_output=True, check=False)
    except Exception as exc:
        return [{"sheet": "", "read_status": "unreadable", "error": str(exc)[:250]}]
    if proc.returncode != 0:
        return [{"sheet": "", "read_status": "unreadable", "error": (proc.stderr or proc.stdout)[:250]}]
    lines = proc.stdout.splitlines()
    if not lines:
        return [{"sheet": "", "read_status": "unreadable", "error": "foreign reader returned no output"}]
    if lines[0].startswith("ERROR\t"):
        return [{"sheet": "", "read_status": "unreadable", "error": lines[0].split("\t", 1)[1]}]
    try:
        _, rows, cols = lines[0].split("\t")
        col_details = []
        for line in lines[1:]:
            parts = line.split("\t")
            if len(parts) < 4:
                continue
            col_details.append({"name": parts[0], "missing": float(parts[1]), "unique": int(float(parts[2])), "sample": parts[3]})
        return [summarize_columns(path.name, path.suffix.lower().lstrip("."), int(rows), int(cols), col_details)]
    except Exception as exc:
        return [{"sheet": "", "read_status": "unreadable", "error": f"{exc}: {proc.stdout[:180]}"}]


def likely_columns(cols: list[str], patterns: list[str]) -> list[str]:
    out = []
    for col in cols:
        low = col.lower()
        if any(re.search(p, low) for p in patterns):
            out.append(col)
    return out[:20]


def summarize_df(df: pd.DataFrame, file_name: str, sheet: str, sampled_rows: int) -> dict[str, object]:
    df = df.copy()
    df.columns = make_unique_names([str(c) for c in df.columns])
    cols = [str(c) for c in df.columns]
    col_details = []
    for c in cols[:120]:
        s = df[c]
        non_null = s.dropna()
        col_details.append(
            {
                "name": c,
                "missing": round(float(s.isna().mean()), 3) if len(s) else 1.0,
                "unique": int(non_null.nunique(dropna=True)) if len(non_null) else 0,
                "sample": "|".join(map(str, non_null.astype(str).unique()[:4])),
            }
        )
    return summarize_columns(file_name, sheet, len(df), len(cols), col_details, sampled_rows=sampled_rows)


def summarize_columns(
    file_name: str,
    sheet: str,
    rows: int,
    cols: int,
    col_details: list[dict[str, object]],
    sampled_rows: int | None = None,
) -> dict[str, object]:
    names = [str(c["name"]) for c in col_details]
    treatment = likely_columns(names, [r"\\bgroup\\b", r"\\barm\\b", r"treat", r"intervention", r"control", r"random", r"alert"])
    outcomes = likely_columns(
        names,
        [
            r"outcome", r"score", r"scale", r"pain", r"death", r"mort", r"los", r"length",
            r"recovery", r"readmission", r"depress", r"hamilton", r"hdrs", r"change",
            r"post", r"follow", r"primary", r"full", r"feed", r"symptom", r"scr",
        ],
    )
    baseline = likely_columns(names, [r"baseline", r"pre", r"age", r"sex", r"gender", r"bmi", r"weight", r"height"])
    id_cols = likely_columns(names, [r"\\bid\\b", r"patient", r"participant", r"subject", r"record"])
    return {
        "sheet": sheet,
        "read_status": "readable",
        "rows": int(rows),
        "cols": int(cols),
        "sampled_rows": int(sampled_rows if sampled_rows is not None else rows),
        "candidate_treatment_columns": "; ".join(treatment),
        "candidate_outcome_columns": "; ".join(outcomes),
        "candidate_baseline_columns": "; ".join(baseline),
        "candidate_id_columns": "; ".join(id_cols),
        "columns_json": json.dumps(col_details[:80], ensure_ascii=False),
        "error": "",
    }


def inspect_table(path: Path) -> list[dict[str, object]]:
    ext = path.suffix.lower()
    if ext in {".csv", ".tsv", ".tab"}:
        return read_delimited(path)
    if ext == ".txt":
        # Treat tabular txt files as candidates; non-tabular support docs will
        # simply show a one-column preview.
        return read_delimited(path)
    if ext == ".xlsx":
        return read_xlsx(path)
    if ext in {".sav", ".dta"}:
        return read_foreign(path)
    if path.name == "RCRC-00890":
        return read_foreign(path)
    if ext in {".xls", ".ods"}:
        return read_excel_with_pandas(path)
    return []


def main() -> int:
    metadata = load_candidate_metadata()
    paths = copy_or_extract_sources()
    file_rows = []
    table_rows = []
    for path in paths:
        cid = candidate_id_from_text(str(path.relative_to(MANUAL_DIR))) if path.is_relative_to(MANUAL_DIR) else candidate_id_from_text(str(path))
        cid = cid or "UNMAPPED"
        rel = safe_rel(path)
        file_rows.append(
            {
                "candidate_id": cid,
                "path": rel,
                "file_name": path.name,
                "extension": path.suffix.lower(),
                "bytes": path.stat().st_size,
                "role": file_role(path),
            }
        )
        for summary in inspect_table(path):
            table_rows.append({"candidate_id": cid, "path": rel, "file_name": path.name, **summary})

    file_df = pd.DataFrame(file_rows).sort_values(["candidate_id", "path"])
    table_df = pd.DataFrame(table_rows).sort_values(["candidate_id", "path", "sheet"])
    decisions = []
    for cid, group in file_df.groupby("candidate_id"):
        tables = table_df[table_df["candidate_id"] == cid]
        readable = tables[tables["read_status"] == "readable"]
        data_files = group[group["role"] == "data_or_table"]
        meta = metadata.get(cid, {})
        title = meta.get("title") or meta.get("Title") or meta.get("dataset_title") or ""
        doi = meta.get("doi") or meta.get("DOI") or ""
        repository = meta.get("repository") or meta.get("Repository") or ""
        has_treatment = any(str(x).strip() for x in readable.get("candidate_treatment_columns", []))
        has_outcome = any(str(x).strip() for x in readable.get("candidate_outcome_columns", []))
        has_participant_rows = any(pd.to_numeric(readable.get("rows", pd.Series(dtype=int)), errors="coerce").fillna(0) >= 20)
        ext_counts = Counter(group["extension"])
        if cid == "UNMAPPED":
            decision = "needs_mapping"
            reason = "No RCTC identifier found in path or filename."
        elif not len(data_files):
            decision = "not_qualified_before_cleaning"
            reason = "No tabular data file found."
        elif not len(readable):
            decision = "needs_reader_or_manual_review"
            reason = "Downloaded tabular files exist but no local reader could inspect them."
        elif not has_participant_rows:
            decision = "not_qualified_before_cleaning"
            reason = "Readable tables do not appear participant-level."
        elif has_treatment and has_outcome:
            decision = "qualified_for_publication_review"
            reason = "Participant-level rows, likely treatment assignment, and likely outcome fields detected."
        else:
            decision = "needs_manual_review_before_cleaning"
            missing = []
            if not has_treatment:
                missing.append("treatment assignment not confirmed")
            if not has_outcome:
                missing.append("outcomes not confirmed")
            reason = "; ".join(missing) or "Ambiguous table structure."
        decisions.append(
            {
                "candidate_id": cid,
                "title": title,
                "doi": doi,
                "repository": repository,
                "file_count": len(group),
                "data_file_count": len(data_files),
                "readable_table_count": len(readable),
                "extensions": "; ".join(f"{k}:{v}" for k, v in sorted(ext_counts.items())),
                "qualification_decision": decision,
                "reasons": reason,
                "best_readable_tables": " | ".join(
                    f"{r.file_name}:{r.sheet} rows={r.rows} cols={r.cols}"
                    for r in readable.itertuples()
                    if int(getattr(r, "rows", 0) or 0) >= 20
                )[:1200],
                "treatment_evidence": "; ".join(sorted(set(filter(None, map(str, readable.get("candidate_treatment_columns", []))))))[:600],
                "outcome_evidence": "; ".join(sorted(set(filter(None, map(str, readable.get("candidate_outcome_columns", []))))))[:600],
            }
        )

    decisions_df = pd.DataFrame(decisions).sort_values("candidate_id")
    file_df.to_csv(WORKSPACE / "manual_file_manifest.csv", index=False)
    table_df.to_csv(WORKSPACE / "manual_table_inspection.csv", index=False)
    decisions_df.to_csv(WORKSPACE / "manual_candidate_decisions_prepub.csv", index=False)
    summary = {
        "candidate_count": int(decisions_df.shape[0]),
        "file_count": int(file_df.shape[0]),
        "readable_table_count": int((table_df["read_status"] == "readable").sum()) if not table_df.empty else 0,
        "decision_counts": decisions_df["qualification_decision"].value_counts().to_dict(),
    }
    (WORKSPACE / "manual_screen_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
