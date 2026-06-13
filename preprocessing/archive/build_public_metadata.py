#!/usr/bin/env python3
"""Build the public 125-trial metadata workbook.

The public workbook keeps the original 19-column Sheet1 contract and appends
the expansion provenance/audit sheets used during curation.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "meta_data.xlsx"
ORIGINAL_CANDIDATES = [
    ROOT / "local" / "legacy_cleaned_data" / "meta_data_original_trials1_50.xlsx",
    ROOT / "meta_data.xlsx",
    ROOT / "cleaned_data" / "meta_data.xlsx",
]
EXPANSION_CANDIDATES = [
    ROOT / "local" / "rct_expansion" / "metadata" / "meta_data_active.xlsx",
    ROOT / "rct_expansion" / "metadata" / "meta_data_active.xlsx",
]

MAIN_COLUMNS = [
    "Trial_ID",
    "Trial Number/Name",
    "Paper Name",
    "Journal",
    "Paper Link",
    "Publication Year",
    "# of Arm",
    "Control Group",
    "Study Phase",
    "Sample Size",
    "Priamry Outcome",
    "Primary Outcome Type",
    "Trial Success(Primary Outcome Significant)",
    "Statistical Model",
    "Randomization Scheme",
    "Randomization Scheme(High Level)",
    "Research Area",
    "Text Data",
    "Citation",
]

COPY_EXPANSION_SHEETS = [
    "Provenance",
    "Validation",
    "Audit_Summary",
    "Audit_Detail",
    "Duplicate_Screen",
    "Cleanup_Issues",
]


def first_existing(paths: list[Path]) -> Path:
    for path in paths:
        if path.exists():
            return path
    raise FileNotFoundError("None of these paths exist: " + ", ".join(map(str, paths)))


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    while rows and all(value is None for value in rows[-1]):
        rows.pop()
    return rows


def rows_as_dicts(ws):
    rows = sheet_rows(ws)
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    out = []
    for row in rows[1:]:
        if all(value is None for value in row):
            continue
        out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return out


def style_sheet(ws):
    if ws.max_row >= 1:
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    for col_idx, column in enumerate(ws.iter_cols(), start=1):
        max_len = 0
        for cell in column[:200]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)


def append_rows(ws, rows):
    for row in rows:
        ws.append(list(row))


def copy_sheet(src_ws, dest_wb, title: str):
    dest = dest_wb.create_sheet(title)
    for row in sheet_rows(src_ws):
        dest.append(list(row))
    style_sheet(dest)


def main():
    original_path = first_existing(ORIGINAL_CANDIDATES)
    expansion_path = first_existing(EXPANSION_CANDIDATES)

    original_wb = load_workbook(original_path, data_only=False)
    expansion_wb = load_workbook(expansion_path, data_only=False)

    original_rows = [
        row for row in rows_as_dicts(original_wb[original_wb.sheetnames[0]])
        if row.get("Trial_ID") is not None and int(row.get("Trial_ID")) <= 50
    ]
    expansion_rows = [
        row for row in rows_as_dicts(expansion_wb["Sheet1"])
        if row.get("Trial_ID") is not None and 51 <= int(row.get("Trial_ID")) <= 125
    ]

    combined = []
    for row in original_rows + expansion_rows:
        values = [row.get(column) for column in MAIN_COLUMNS]
        values[0] = int(values[0])
        combined.append(values)

    trial_ids = [int(row[0]) for row in combined if row[0] is not None]
    if sorted(trial_ids) != list(range(1, 126)):
        raise ValueError(f"Expected Trial_ID 1:125 exactly, found {min(trial_ids)}:{max(trial_ids)} ({len(trial_ids)} rows)")
    if len(set(trial_ids)) != 125:
        raise ValueError("Duplicate Trial_ID values found in combined metadata.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    append_rows(ws, [MAIN_COLUMNS] + combined)
    style_sheet(ws)

    for sheet_name in COPY_EXPANSION_SHEETS:
        if sheet_name in expansion_wb.sheetnames:
            copy_sheet(expansion_wb[sheet_name], wb, sheet_name)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Sheet1 rows: {len(combined)}")
    print(f"Expansion metadata source: {expansion_path}")
    print(f"Original metadata source: {original_path}")


if __name__ == "__main__":
    main()
