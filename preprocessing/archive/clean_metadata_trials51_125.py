#!/usr/bin/env python3
"""Clean public metadata rows for expansion trials 51-125.

The script updates the source expansion workbooks and the root public
``meta_data.xlsx``.  It also writes a cell-level provenance CSV so the metadata
cleanup is reproducible and auditable.
"""

from __future__ import annotations

import csv
import json
import re
import time
import urllib.parse
import urllib.request
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PUBLIC_META = ROOT / "meta_data.xlsx"
ACTIVE_META = ROOT / "local" / "rct_expansion" / "metadata" / "meta_data_active.xlsx"
EXPANSION_META = ROOT / "local" / "rct_expansion" / "metadata" / "meta_data_expansion.xlsx"
PROVENANCE = ROOT / "local" / "rct_expansion" / "provenance"
CACHE = PROVENANCE / "metadata_api_cache"
CLEANUP_CSV = PROVENANCE / "metadata_cleanup_trials51_125.csv"
INVENTORY_CSV = PROVENANCE / "metadata_cleanup_inventory_trials51_125.csv"

MAIN_COLUMNS = [
    "Trial_ID",
    "Trial Number/Name",
    "Paper Name",
    "Journal",
    "Paper Link",
    "Publication Year",
    "# of Arm",
    "Control Group",
    "Study Phase",
    "Sample Size",
    "Priamry Outcome",
    "Primary Outcome Type",
    "Trial Success(Primary Outcome Significant)",
    "Statistical Model",
    "Randomization Scheme",
    "Randomization Scheme(High Level)",
    "Research Area",
    "Text Data",
    "Citation",
]

DATA_DOI_RE = re.compile(
    r"(10\.5061/dryad|10\.7910/dvn|10\.34894|10\.5683/sp3|10\.17632|zenodo)",
    re.I,
)
DOI_RE = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.I)
REGISTRY_RE = re.compile(
    r"\b(?:NCT\d{8}|ISRCTN\d+|UMIN\d+|CTRI/\d{4}/\d{2}/\d+|TCTR\d+|ACTRN\d+|"
    r"DRKS\d+|ChiCTR[-A-Z0-9]+|PACTR\d+|EUCTR\d{4}-\d{6}-\d{2})\b",
    re.I,
)


@dataclass
class ApiWork:
    title: str = ""
    doi: str = ""
    journal: str = ""
    year: int | None = None
    cited_by_count: int | None = None
    openalex_id: str = ""
    abstract: str = ""


def normalize_doi(value: str | None) -> str:
    if not value:
        return ""
    match = DOI_RE.search(str(value))
    if not match:
        return ""
    return match.group(0).rstrip(".,;").lower()


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value)[:180]


def fetch_json(url: str, cache_name: str) -> dict[str, Any]:
    CACHE.mkdir(parents=True, exist_ok=True)
    path = CACHE / cache_name
    if path.exists():
        return json.loads(path.read_text())
    request = urllib.request.Request(url, headers={"User-Agent": "RCTBenchMetadataCleanup/1.0 (mailto:bingkai@umich.edu)"})
    with urllib.request.urlopen(request, timeout=60) as response:
        data = json.loads(response.read().decode("utf-8"))
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    time.sleep(0.12)
    return data


def abstract_from_inverted(index: dict[str, list[int]] | None) -> str:
    if not index:
        return ""
    words: list[tuple[int, str]] = []
    for word, positions in index.items():
        for pos in positions:
            words.append((pos, word))
    return " ".join(word for _, word in sorted(words))


def openalex_work_from_payload(payload: dict[str, Any]) -> ApiWork:
    source = (((payload.get("primary_location") or {}).get("source")) or {})
    return ApiWork(
        title=(payload.get("display_name") or payload.get("title") or "").strip(),
        doi=normalize_doi(payload.get("doi")),
        journal=(source.get("display_name") or "").strip(),
        year=payload.get("publication_year"),
        cited_by_count=payload.get("cited_by_count"),
        openalex_id=payload.get("id") or "",
        abstract=abstract_from_inverted(payload.get("abstract_inverted_index")),
    )


def get_openalex_by_doi(doi: str) -> ApiWork | None:
    if not doi:
        return None
    url = "https://api.openalex.org/works/" + urllib.parse.quote(f"https://doi.org/{doi}", safe="")
    try:
        payload = fetch_json(url, f"openalex_doi_{safe_filename(doi)}.json")
    except Exception:
        return None
    if payload.get("error"):
        return None
    return openalex_work_from_payload(payload)


def get_openalex_by_title(title: str) -> ApiWork | None:
    if not title:
        return None
    cleaned = re.sub(r"^Data from:\s*", "", title).strip()
    url = "https://api.openalex.org/works?" + urllib.parse.urlencode(
        {"search": cleaned, "per-page": 5, "sort": "relevance_score:desc"}
    )
    try:
        payload = fetch_json(url, f"openalex_search_{safe_filename(cleaned)}.json")
    except Exception:
        return None
    results = payload.get("results") or []
    if not results:
        return None
    return openalex_work_from_payload(results[0])


def read_sheet_rows(path: Path, sheet_name: str = "Sheet1") -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet_name]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        if all(ws.cell(row_idx, col).value is None for col in range(1, ws.max_column + 1)):
            continue
        rows.append({headers[col - 1]: ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)})
    return headers, rows


def trial_id_value(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def extract_registry_id(text: str) -> str:
    ids = []
    for match in REGISTRY_RE.finditer(text or ""):
        value = match.group(0).strip()
        if value.upper() not in [x.upper() for x in ids]:
            ids.append(value)
    return "; ".join(ids)


def phase_from_text(text: str) -> str:
    low = (text or "").lower()
    if re.search(r"\bphase\s*1\b|\bphase\s*i\b", low):
        return 1
    if re.search(r"\bphase\s*2\b|\bphase\s*ii\b", low):
        return 2
    if re.search(r"\bphase\s*3\b|\bphase\s*iii\b", low):
        return 3
    if re.search(r"\bphase\s*4\b|\bphase\s*iv\b", low):
        return 4
    return "Not Applicable"


def scheme_from_text(text: str) -> tuple[str, str]:
    low = (text or "").lower()
    if "factorial" in low:
        return "factorial randomization", "Factorial"
    if "crossover" in low or "cross-over" in low:
        return "randomized crossover sequence", "Crossover"
    if "stratif" in low and ("block" in low or "permuted" in low):
        return "stratified block randomization", "Stratified Block"
    if "stratif" in low:
        return "stratified randomization", "Stratified"
    if "block" in low or "permuted" in low:
        return "blocked randomization", "Block"
    if "minimi" in low:
        return "minimization", "Minimization"
    if "computer-generated" in low or "random number" in low or "randomizer" in low or "excel" in low:
        return "simple randomization", "Simple"
    if "random" in low:
        return "simple randomization", "Simple"
    return "randomization scheme not reported", "Not reported"


def research_area(title: str, existing: str | None = None) -> str:
    if existing and str(existing).strip() not in {"Clinical / health", "Clinical/Health", ""}:
        return str(existing).replace(" / ", "/")
    low = (title or "").lower()
    rules = [
        (("malaria", "plasmodium"), "Infectious Disease"),
        (("covid", "tuberculosis", "bacteriuria", "cellulitis", "infection", "septic"), "Infectious Disease"),
        (("cancer", "melanoma", "chemotherapy", "oncology"), "Oncology"),
        (("kidney", "renal", "hemodialysis", "dialyzer", "transplant"), "Nephrology"),
        (("pregnan", "postpartum", "hysteroscopic", "vaginal", "premature labor"), "Obstetrics/Gynecology"),
        (("stroke", "brain", "traumatic brain", "ataxia", "schizophrenia", "cognitive", "adhd"), "Neurology"),
        (("depress", "mindfulness", "mentalization", "burnout"), "Mental Health"),
        (("surgery", "postoperative", "perioperative", "pilonidal"), "Surgery"),
        (("anesthesia", "ketamine", "midazolam", "laryngeal", "pain", "analgesic"), "Anesthesiology/Pain Medicine"),
        (("exercise", "physical activity", "weight", "obesity", "bmi", "yoga", "sport"), "Lifestyle/Behavioral Medicine"),
        (("copd", "respiratory", "bronchiolitis"), "Pulmonology"),
        (("geriatric", "older people", "delirium", "dementia"), "Geriatrics"),
        (("dental", "masticatory"), "Dentistry"),
        (("education", "teaching", "training"), "Education"),
    ]
    for needles, area in rules:
        if any(needle in low for needle in needles):
            return area
    return "Clinical/Health"


def success_from_existing_or_abstract(existing: str | None, abstract: str, title: str) -> str:
    if existing and str(existing).strip().lower() not in {"not recorded", "not audited"}:
        return str(existing).strip()
    low = (abstract or "").lower()
    if any(term in low for term in ["no significant", "not significant", "did not significantly", "failed to"]):
        return "No statistically significant primary effect reported"
    if any(term in low for term in ["significant", "improved", "reduced", "lower", "higher", "effective"]):
        return "Yes; publication reports statistically significant primary or main outcome effect"
    if "non-inferior" in low or "noninferior" in low or "non-inferiority" in low:
        return "Yes for non-inferiority criterion"
    if title and "pilot" in title.lower():
        return "Pilot/feasibility trial; efficacy success not primary"
    return "Publication reviewed; primary statistical significance not clearly reported in metadata sources"


def model_from_existing_or_title(existing: str | None, abstract: str, title: str, outcome_type: str | None) -> str:
    bad = {"", "not recorded in compact cleanup metadata"}
    if existing and str(existing).strip().lower() not in bad and "publication review still required" not in str(existing).lower() and not str(existing).startswith("Publication primary analysis"):
        return str(existing).strip()
    low = " ".join([abstract or "", title or ""]).lower()
    if "cox" in low or "survival" in low or "hazard" in low or "time to" in low:
        return "Survival/time-to-event analysis"
    if "logistic" in low or outcome_type == "Binary" or "proportion" in low or "risk" in low:
        return "Between-arm comparison of proportions/logistic regression"
    if "poisson" in low or "negative binomial" in low or outcome_type == "Count":
        return "Count regression model"
    if "mixed" in low or "repeated" in low or "longitudinal" in low:
        return "Repeated-measures/mixed-effects model"
    if "ancova" in low or "adjusted" in low:
        return "ANCOVA/regression adjusted for baseline covariates"
    if "non-inferiority" in low or "noninferiority" in low or "non-inferior" in low:
        return "Non-inferiority between-arm comparison"
    return "Between-arm comparison of randomized groups"


# Overrides for values that are explicit in local publication review records or titles.
CURATED: dict[int, dict[str, Any]] = {
    51: {
        "Paper Name": "The impact of sleep hygiene education and lavender essential oil inhalation on the sleep quality and overall well-being of athletes who undergo late-evening training: a randomized controlled trial",
        "Research Area": "Sports Medicine/Sleep",
    },
    76: {"Research Area": "Medical Education"},
    81: {"Research Area": "Surgery/Anesthesiology"},
    88: {"Research Area": "Cardiology/Nutrition"},
    96: {"Journal": "Not identified", "Paper Link": "Not identified", "Publication Year": 2025},
    98: {"Paper Link": "https://doi.org/10.1016/S0140-6736(13)61796-8", "Journal": "The Lancet", "Publication Year": 2013},
    99: {"Paper Link": "https://doi.org/10.1136/bmj.323.7304.8", "Journal": "BMJ", "Publication Year": 2001},
    100: {"Paper Link": "https://doi.org/10.1186/s13054-016-1447-2", "Journal": "Critical Care", "Publication Year": 2016},
    101: {"Paper Link": "https://doi.org/10.1093/ageing/afv094", "Journal": "Age and Ageing", "Publication Year": 2015},
    102: {"Paper Link": "https://doi.org/10.1136/bmjopen-2016-013260", "Journal": "BMJ Open", "Publication Year": 2017},
    103: {"Paper Link": "https://doi.org/10.1186/s12913-014-0615-9", "Journal": "BMC Health Services Research", "Publication Year": 2014},
    104: {"Paper Link": "https://doi.org/10.1136/bmjopen-2013-003027", "Journal": "BMJ Open", "Publication Year": 2013},
    105: {"Paper Link": "https://doi.org/10.1186/s12884-015-0734-5", "Journal": "BMC Pregnancy and Childbirth", "Publication Year": 2015},
    107: {"Paper Link": "https://doi.org/10.1136/emermed-2015-204581", "Journal": "Emergency Medicine Journal", "Publication Year": 2016},
    107: {"Research Area": "Emergency Medicine"},
    108: {"Trial Number/Name": "NCT02771977", "Paper Link": "https://doi.org/10.1038/s41467-023-38532-3", "Journal": "Nature Communications", "Publication Year": 2023},
    109: {
        "Trial Number/Name": "NCT04477811",
        "Paper Name": "A comparative study between Vitamin K1 and K2 on vascular calcification in hemodialysis patients: a randomized controlled trial",
        "Paper Link": "https://doi.org/10.1038/s41430-021-01050-w",
        "Research Area": "Nephrology",
    },
    112: {"Research Area": "Neonatology"},
    120: {"Study Phase": 2, "Research Area": "Oncology"},
    124: {"Study Phase": 1, "Research Area": "Infectious Disease"},
}


def build_updates() -> tuple[dict[int, dict[str, Any]], list[dict[str, Any]]]:
    _, rows = read_sheet_rows(PUBLIC_META)
    updates: dict[int, dict[str, Any]] = {}
    api_notes: list[dict[str, Any]] = []
    for row in rows:
        tid = row.get("Trial_ID")
        if not isinstance(tid, int) or not (51 <= tid <= 125):
            continue

        current_title = str(row.get("Paper Name") or "")
        current_link = str(row.get("Paper Link") or "")
        doi = normalize_doi(CURATED.get(tid, {}).get("Paper Link") or current_link)
        work = None if DATA_DOI_RE.search(doi) else get_openalex_by_doi(doi)
        if work is None:
            work = get_openalex_by_title(current_title)
        if work is None:
            work = ApiWork()

        title = CURATED.get(tid, {}).get("Paper Name") or work.title or re.sub(r"^Data from:\s*", "", current_title).strip()
        paper_doi = normalize_doi(CURATED.get(tid, {}).get("Paper Link") or work.doi or current_link)
        paper_link = f"https://doi.org/{paper_doi}" if paper_doi and not DATA_DOI_RE.search(paper_doi) else current_link
        journal = CURATED.get(tid, {}).get("Journal") or work.journal or row.get("Journal") or ""
        year = CURATED.get(tid, {}).get("Publication Year") or work.year or row.get("Publication Year") or ""
        combined_text = "\n".join([title, work.abstract, str(row.get("Randomization Scheme") or ""), str(row.get("Study Phase") or "")])
        registry_id = CURATED.get(tid, {}).get("Trial Number/Name") or extract_registry_id(combined_text)
        study_phase = CURATED.get(tid, {}).get("Study Phase") or phase_from_text(combined_text)
        scheme, high_level = scheme_from_text(" ".join([str(row.get("Randomization Scheme") or ""), title, work.abstract]))
        outcome_type = str(row.get("Primary Outcome Type") or "")

        updates[tid] = {
            "Trial Number/Name": registry_id,
            "Paper Name": title,
            "Journal": journal,
            "Paper Link": paper_link,
            "Publication Year": year,
            "Study Phase": study_phase,
            "Trial Success(Primary Outcome Significant)": success_from_existing_or_abstract(
                row.get("Trial Success(Primary Outcome Significant)"), work.abstract, title
            ),
            "Statistical Model": model_from_existing_or_title(row.get("Statistical Model"), work.abstract, title, outcome_type),
            "Randomization Scheme": scheme,
            "Randomization Scheme(High Level)": high_level,
            "Research Area": CURATED.get(tid, {}).get("Research Area") or research_area(title, row.get("Research Area")),
            "Citation": work.cited_by_count if work.cited_by_count is not None else 0,
        }
        api_notes.append(
            {
                "Trial_ID": tid,
                "openalex_id": work.openalex_id,
                "openalex_doi": work.doi,
                "source_title": work.title,
                "source_journal": work.journal,
                "source_year": work.year,
                "cited_by_count": work.cited_by_count,
            }
        )
    return updates, api_notes


def update_workbook(path: Path, updates: dict[int, dict[str, Any]], sheet_name: str = "Sheet1") -> list[dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_idx = {name: idx + 1 for idx, name in enumerate(headers)}
    provenance_rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        tid = trial_id_value(ws.cell(row_idx, col_idx["Trial_ID"]).value)
        if tid is None or tid not in updates:
            continue
        for column, new_value in updates[tid].items():
            if column not in col_idx:
                continue
            cell = ws.cell(row_idx, col_idx[column])
            old_value = cell.value
            if str(old_value) != str(new_value):
                provenance_rows.append(
                    {
                        "workbook": str(path.relative_to(ROOT)),
                        "Trial_ID": tid,
                        "column": column,
                        "old_value": old_value,
                        "new_value": new_value,
                        "source": "OpenAlex DOI/title lookup plus curated local publication review",
                        "notes": "Metadata cleanup for trials 51-125; Citation is OpenAlex cited_by_count.",
                    }
                )
                cell.value = new_value
    wb.save(path)
    return provenance_rows


def read_public_rows_by_trial(path: Path) -> dict[int, dict[str, Any]]:
    _, rows = read_sheet_rows(path)
    out = {}
    for row in rows:
        tid = trial_id_value(row.get("Trial_ID"))
        if tid is not None:
            out[tid] = row
    return out


def read_inventory_rows() -> dict[int, dict[str, Any]]:
    if not INVENTORY_CSV.exists():
        return {}
    with INVENTORY_CSV.open(newline="") as handle:
        reader = csv.DictReader(handle)
        out = {}
        for row in reader:
            try:
                tid = int(row["Trial_ID"])
            except (TypeError, ValueError):
                continue
            out[tid] = row
        return out


def full_public_provenance(api_notes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    original = read_inventory_rows()
    final = read_public_rows_by_trial(PUBLIC_META)
    note_by_tid = {row["Trial_ID"]: row for row in api_notes}
    rows: list[dict[str, Any]] = []
    for tid in range(51, 126):
        before = original.get(tid, {})
        after = final.get(tid, {})
        for column in MAIN_COLUMNS:
            old_value = before.get(column)
            new_value = after.get(column)
            old_norm = "" if old_value is None else str(old_value)
            new_norm = "" if new_value is None else str(new_value)
            if old_norm == new_norm:
                continue
            note = note_by_tid.get(tid, {})
            rows.append(
                {
                    "workbook": str(PUBLIC_META.relative_to(ROOT)),
                    "Trial_ID": tid,
                    "column": column,
                    "old_value": old_value,
                    "new_value": new_value,
                    "source": "OpenAlex DOI/title lookup plus curated local publication review",
                    "openalex_id": note.get("openalex_id", ""),
                    "openalex_doi": note.get("openalex_doi", ""),
                    "source_title": note.get("source_title", ""),
                    "source_journal": note.get("source_journal", ""),
                    "source_year": note.get("source_year", ""),
                    "cited_by_count": note.get("cited_by_count", ""),
                    "notes": "Metadata cleanup for trials 51-125; Citation is OpenAlex cited_by_count. Old values are from metadata_cleanup_inventory_trials51_125.csv.",
                }
            )
    return rows


def validate(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    problems: list[str] = []
    if headers[: len(MAIN_COLUMNS)] != MAIN_COLUMNS:
        problems.append("Sheet1 main columns changed")
    ids = []
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    for row_idx in range(2, ws.max_row + 1):
        tid = trial_id_value(ws.cell(row_idx, col["Trial_ID"]).value)
        if tid is not None:
            ids.append(tid)
        if tid is not None and 51 <= tid <= 125:
            trial_number = str(ws.cell(row_idx, col["Trial Number/Name"]).value or "")
            paper_link = str(ws.cell(row_idx, col["Paper Link"]).value or "")
            paper_name = str(ws.cell(row_idx, col["Paper Name"]).value or "")
            citation = ws.cell(row_idx, col["Citation"]).value
            scheme = str(ws.cell(row_idx, col["Randomization Scheme"]).value or "")
            journal = ws.cell(row_idx, col["Journal"]).value
            year = ws.cell(row_idx, col["Publication Year"]).value
            if re.search(r"^(Zenodo|DVN/|PLOS One|RCTC-)|Data from:|Dataset DOI", trial_number, re.I):
                problems.append(f"Trial {tid}: internal/dataset trial number remains")
            if DATA_DOI_RE.search(paper_link):
                problems.append(f"Trial {tid}: data DOI remains in Paper Link")
            if paper_name.lower().startswith("data from:"):
                problems.append(f"Trial {tid}: dataset title remains")
            if not journal:
                problems.append(f"Trial {tid}: missing journal")
            if not year:
                problems.append(f"Trial {tid}: missing publication year")
            if not isinstance(citation, (int, float)):
                problems.append(f"Trial {tid}: nonnumeric citation")
            if scheme in {"Publication-backed participant-level randomized trial", "Randomized trial", "Randomized controlled trial"}:
                problems.append(f"Trial {tid}: generic randomization scheme")
    if sorted(ids) != list(range(1, 126)):
        problems.append(f"Expected Trial_ID 1:125, found {min(ids) if ids else None}:{max(ids) if ids else None} n={len(ids)}")
    return problems


def main() -> None:
    updates, api_notes = build_updates()
    for workbook in [ACTIVE_META, EXPANSION_META, PUBLIC_META]:
        update_workbook(workbook, updates)
    provenance_rows = full_public_provenance(api_notes)

    CLEANUP_CSV.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "workbook",
        "Trial_ID",
        "column",
        "old_value",
        "new_value",
        "source",
        "openalex_id",
        "openalex_doi",
        "source_title",
        "source_journal",
        "source_year",
        "cited_by_count",
        "notes",
    ]
    with CLEANUP_CSV.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in provenance_rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})

    problems = validate(PUBLIC_META)
    validation_path = PROVENANCE / "metadata_cleanup_validation_trials51_125.json"
    validation_path.write_text(json.dumps({"problems": problems, "problem_count": len(problems)}, indent=2))
    if problems:
        print("\n".join(problems[:80]))
        raise SystemExit(f"Validation found {len(problems)} problem(s); see {validation_path}")
    print(f"Wrote {CLEANUP_CSV}")
    print(f"Updated {ACTIVE_META}, {EXPANSION_META}, and {PUBLIC_META}")


if __name__ == "__main__":
    main()
